import json
import time
from operator import itemgetter

import openpyxl
import requests

start_time = time.time()

wb = openpyxl.load_workbook('report.xlsx')

YDAuth = {'za4aroff.ol': 'AgAAAAAzN9ktAAXAInwzJfwlY0E4taHqCQUC33I',
          'macsim.reshetar': 'AgAAAAArGY4IAAXAIq_OsOIJXEgIvBjrf0dpP0s'}

direct_client_logins = ["za4aroff.ol", "za4aroff.ol", "macsim.reshetar"]  # НП СПБ, НП МСК, 1001
clientTokens = {"za4aroff.ol": "AgAAAAAzN9ktAAXAInwzJfwlY0E4taHqCQUC33I",
                "macsim.reshetar": "AgAAAAArGY4IAAXAIq_OsOIJXEgIvBjrf0dpP0s"}
date1 = "2019-09-01"
date2 = "2019-09-23"
ids = ["35656455", "37340325", "42890799"]  # НП СПБ, НП МСК, 1001
metrics = ["ym:s:goal39547444visits", "ym:s:goal39545641visits", "ym:s:goal44259842visits"]  # НП СПБ, НП МСК, 1001
dimensions = "ym:s:goal,ym:s:lastsignDirectClickOrder,ym:s:lastsignDirectBannerGroup"
filters = ["ym:s:goal==39547444", "ym:s:goal==39545641", "ym:s:goal==44259842"]
url = "https://api-metrika.yandex.net/stat/v1/data"


def saveConversions(login_list, clientTokens, directIDs, metrics, startDate, endDate, filters):
    if 'Конверсии' in wb.sheetnames:
        report = wb['Конверсии']
    else:
        report = wb.create_sheet('Конверсии')

    for n, field in enumerate(['Кампания', 'Группа', 'ID Группы', 'Конверсии']):  # Вводим поля таблицы
        _ = report.cell(column=1 + n, row=1, value=field)

    for n, acc in enumerate(login_list):
        requestURL = url + "?direct_client_logins=%s&ids=%s&metrics=%s&date1=%s&date2=%s&dimensions=%s&filters=%s" \
                           "&pretty=true" % (acc, directIDs[n], metrics[n], startDate, endDate, dimensions, filters[n])

        headers = {"Authorization": "OAuth " + clientTokens.get(acc), "Client-Login": acc, }
        req = requests.get(requestURL, headers=headers)
        # print(req.text)
        jsonReport = json.loads(req.text)

        maxRow = report.max_row
        print(jsonReport["data"])
        for index, item in enumerate(jsonReport["data"]):
            if item['dimensions'][1]['name'] != "Other campaigns":
                _ = report.cell(row=index + maxRow, column=1, value=item['dimensions'][1]['name'])
                _ = report.cell(row=index + maxRow, column=2, value=item['dimensions'][2]['name'])
                _ = report.cell(row=index + maxRow, column=3, value=item['dimensions'][2]['id'])
                _ = report.cell(row=index + maxRow, column=4, value=int(item['metrics'][0]))
                print("Найдены конверсии в кампании %s, в группе %s в количестве: %s." % \
                      (item['dimensions'][1]['name'], item['dimensions'][2]['name'], int(item['metrics'][0])))
    wb.save(filename="report.xlsx")


def connectConversionsWithCampaigns():  # Добавляем конверсии в таблицу с данными
    data = wb["Данные"]
    conversions = wb["Конверсии"]
    for n, field in enumerate(["Конверсии", "% конверсии", "CPA"]):  # Вводим поля таблицы
        _ = data.cell(column=8 + n, row=1, value=field)

    dataList = []  # Сортируем "Данные" по группе объявлений
    for n, row in enumerate(data):
        dataList.append([])
        for col in row:
            dataList[n].append(col.value)
    sortedCosts = sorted(dataList[1:], key=itemgetter(3))
    maxCol = data.max_column
    for n, row in enumerate(sortedCosts):
        for col in range(maxCol):
            _ = data.cell(row=n + 2, column=col + 1, value=row[col])

    dataList = []  # Сортируем "Конверсии" по группе объявлений
    dataRow = 0
    for n, row in enumerate(conversions):
        if row[2].value != None:
            dataList.append([])
            for m, col in enumerate(row):
                # print(col.value)
                if col.value == None:
                    dataList[dataRow].append("")
                elif str(col.value).isdigit():
                    dataList[dataRow].append(int(col.value))
                else:
                    dataList[dataRow].append(str(col.value))
            dataRow += 1

    print(dataList)
    sortedConversions = sorted(dataList[1:], key=itemgetter(2))
    maxCol = conversions.max_column
    for n, row in enumerate(sortedConversions):
        for col in range(maxCol):
            _ = conversions.cell(row=n + 2, column=col + 1, value=row[col])

    convCounter = 1
    for camp in range(1, data.max_row):  # Соотносим конверсии с группами объявлений
        if convCounter < conversions.max_row:
            if data["D"][camp].value == int(conversions["C"][convCounter].value):
                _ = data.cell(row=camp + 1, column=8, value=int(conversions["D"][convCounter].value))
                _ = data.cell(row=camp + 1, column=10, value=float(data["G"][camp].value)
                                                             / int(conversions["D"][convCounter].value))
                if data["F"][camp].value > 0:
                    _ = data.cell(row=camp + 1, column=9,
                                  value=int(conversions["D"][convCounter].value) / data["F"][camp].value)
                else:
                    _ = data.cell(row=camp + 1, column=9, value=0)
                print("Выгружаю конверсии для %s из кампании %s." %
                      (conversions["B"][convCounter].value, conversions["A"][convCounter].value))
                convCounter += 1
        else:
            break

    wb.save(filename="report.xlsx")


def mainFunc(login_list, clientTokens, directIDs, metrics, startDate, endDate, filters):
    saveConversions(login_list, clientTokens, directIDs, metrics, startDate, endDate, filters)
    connectConversionsWithCampaigns()


mainFunc(direct_client_logins, clientTokens, ids, metrics, date1, date2, filters)
print("\n--- %s seconds ---" % (time.time() - start_time))
