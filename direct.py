# -*- coding: utf-8 -*-
import json
import sys
from time import sleep

import requests
from openpyxl import Workbook

if sys.version_info < (3,):
    def u(x):
        try:
            return x.encode("utf8")
        except UnicodeDecodeError:
            return x
else:
    def u(x):
        if type(x) == type(b''):
            return x.decode('utf8')
        else:
            return x

# token = 'AgAAAAAJq2baAAXKPEEN2051SEebhC-XYWWB8kQ'  # Токен от zakharoffffff
token = 'AgAAAAArGY4IAAXAIq_OsOIJXEgIvBjrf0dpP0s' # Токен от macsim.reshetar
#token = 'AgAAAAAzN9ktAAXAInwzJfwlY0E4taHqCQUC33I' # Токен от za4aroff.ol
login = 'macsim.reshetar'


class Account(object):

    def __init__(self, token, login):
        self.token = token
        self.login = login

    campaignList = []
    CampaignsURL = 'https://api.direct.yandex.com/json/v5/campaigns'
    ReportsURL = 'https://api.direct.yandex.com/json/v5/reports'

    def getCosts(self):
        self.responseReport = ''
        headers = {
            "Authorization": "Bearer " + self.token,
            "Client-Login": self.login,
            "Accept-Language": "ru",
            "processingMode": "auto"
            # Формат денежных значений в отчете
            # "returnMoneyInMicros": "false",
            # Не выводить в отчете строку с названием отчета и диапазоном дат
            # "skipReportHeader": "true",
            # Не выводить в отчете строку с названиями полей
            # "skipColumnHeader": "true",
            # Не выводить в отчете строку с количеством строк статистики
            # "skipReportSummary": "true"
        }
        self.costsBody = {
            "params": {
                "SelectionCriteria": {
                    "DateFrom": "2019-07-01",
                    "DateTo": "2019-07-31",
                    "Filter": [{
                        "Field": "Impressions",
                        "Operator": "GREATER_THAN",
                        "Values": [0],
                    }],
                },
                "FieldNames": [
                    # "Date",
                    "CampaignId",
                    "CampaignName",
                    # "LocationOfPresenceName",
                    # "Impressions",
                    # "Clicks",
                    "Cost"
                ],
                "ReportName": u("НАЗВАНИЕ_ОТЧЕТА"),
                "ReportType": "CUSTOM_REPORT",
                "DateRangeType": "CUSTOM_DATE",
                "Format": "TSV",
                "IncludeVAT": "YES",
                "IncludeDiscount": "NO"
            }
        }

        # Кодирование тела запроса в JSON
        body = json.dumps(self.costsBody, indent=4)

        # --- Запуск цикла для выполнения запросов ---
        # Если получен HTTP-код 200, то выводится содержание отчета
        # Если получен HTTP-код 201 или 202, выполняются повторные запросы
        while True:
            try:
                req = requests.post(self.ReportsURL, body, headers=headers)
                req.encoding = 'utf-8'  # Принудительная обработка ответа в кодировке UTF-8
                print(req.text)
                if req.status_code == 400:
                    print("Параметры запроса указаны неверно или достигнут лимит отчетов в очереди")
                    print("RequestId: {}".format(req.headers.get("RequestId", False)))
                    print("JSON-код запроса: {}".format(u(body)))
                    print("JSON-код ответа сервера: \n{}".format(u(req.json())))
                    break
                elif req.status_code == 200:
                    # print("RequestId: {}".format(req.headers.get("RequestId", False)))
                    # print("Содержание отчета: \n{}".format(u(req.text)))
                    self.responseReport += req.text
                    print("Данные о расходах кампаний получены")
                    break
                elif req.status_code == 201:
                    print("Отчет успешно поставлен в очередь в режиме офлайн")
                    retryIn = int(req.headers.get("retryIn", 60))
                    print("Повторная отправка запроса через {} секунд".format(retryIn))
                    print("RequestId: {}".format(req.headers.get("RequestId", False)))
                    sleep(retryIn)
                elif req.status_code == 202:
                    print("Отчет формируется в режиме офлайн")
                    retryIn = int(req.headers.get("retryIn", 60))
                    print("Повторная отправка запроса через {} секунд".format(retryIn))
                    print("RequestId:  {}".format(req.headers.get("RequestId", False)))
                    sleep(retryIn)
                elif req.status_code == 500:
                    print("При формировании отчета произошла ошибка. Пожалуйста, попробуйте повторить запрос "
                          "позднее")
                    print("RequestId: {}".format(req.headers.get("RequestId", False)))
                    print("JSON-код ответа сервера: \n{}".format(u(req.json())))
                    break
                elif req.status_code == 502:
                    print("Время формирования отчета превысило серверное ограничение.")
                    print("Пожалуйста, попробуйте изменить параметры запроса - уменьшить период и количество "
                          "запрашиваемых данных.")
                    print("JSON-код запроса: {}".format(body))
                    print("RequestId: {}".format(req.headers.get("RequestId", False)))
                    print("JSON-код ответа сервера: \n{}".format(u(req.json())))
                    break
                else:
                    print("Произошла непредвиденная ошибка")
                    print("RequestId:  {}".format(req.headers.get("RequestId", False)))
                    print("JSON-код запроса: {}".format(body))
                    print("JSON-код ответа сервера: \n{}".format(u(req.json())))
                    break

                    # Обработка ошибки, если не удалось соединиться с сервером API Директа
            except requests.exceptions.ConnectionError:
                # В данном случае мы рекомендуем повторить запрос позднее
                print("Произошла ошибка соединения с сервером API")
                # Принудительный выход из цикла
                break

                # Если возникла какая-либо другая ошибка
            except requests.exceptions.RequestException as e:
                # В данном случае мы рекомендуем проанализировать действия приложения
                print(e)
                print("Произошла непредвиденная ошибка")
                # Принудительный выход из цикла
                break
            # self.campaignCosts = collections.OrderedDict(sorted(self.campaignCosts.items()))
            print(self.responseReport)

    def exportToExcel(self):
        self.getCosts()
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт YD"
        for col in range(len(self.costsBody.get("params").get("FieldNames"))):  # Проставляем графы таблицы
            _ = ws.cell(column=col + 1, row=1, value=self.costsBody.get("params").get("FieldNames")[col])

        rows = self.responseReport.split('\n')
        for m, camp in enumerate(rows[2:len(rows) - 2]):
            campaign = camp.split(sep='\t')
            for n, field in enumerate(campaign):
                if field.isdigit():
                    _ = ws.cell(column=n + 1, row=m + 2, value=int(field))
                else:
                    _ = ws.cell(column=n + 1, row=m + 2, value=field)
        wb.save(filename="report.xlsx")


YDUser = Account(token, login)

YDUser.exportToExcel()
