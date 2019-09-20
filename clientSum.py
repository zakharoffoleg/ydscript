from __future__ import print_function

import time
from operator import itemgetter

import openpyxl

start_time = time.time()

clientList = {
    # Заявки
    "А Эстейт": "А_ЭСТ", "Абсолют Строй Сервис": "АСС_", "Бекар": "БЕКАР", "Вита Ропшинский Квартал": "ВИТА",
    "ВТБ Арена Парк": "ВТБ", "ЖК Полет Тасп Технолоджи": "ТАСП", "Ленстройтрест": "ЛЕНСТР",
    "ЛСР Комфорт": "ЛСР_К", "ЛСР Премиум": "ЛСР_П", "Мастер на Серебристом": "МАСТЕР_Д", "Метрика": "МЕТРИКА",
    "ПАН": "ПАН_", "Патриот Нева": "ПАТРИОТ", "Петрострой": "ПЕТРОСТРОЙ",
    "Рождественно Мир Митино": "РОЖДЕСТВЕНО", "Росстрой Инвест": "РОССТР", "Тренд": "РАЙОНЫ",
    "Хочу Квартиру": "ХОЧУ", "Эталон Серебряный фонтан": "ЭТАЛОН_ИНВЕСТ_ФОНТАН", "Северный город": "СЕВЕРН",
    "ПСК": "ПСК",

    # Спецпроекты
    "Альянс Добрыня - 2": "АЛЬЯНС_ДОБРЫНЯ", "АСД Маленькая Швейцария": "ШВЕЙЦАРИЯ", "Афи Сильвер": "AFI",
    "Балтийская Жемчужина": "БАЛТИЙСКАЯ", "БФА Огни Залива": "БФА", "Вита Федоровское": "ВИТА_Ф",
    "Легенда": "LEGENDA", "ЛенСпецСму бизнес": "ЛЕНСПЕЦСМУ_БИЗНЕС", "ЛенСпецСму комфорт": "ЛЕНСПЕЦСМУ_КОМФОРТ",
    "ЛСР Лучи": "ЛСР_МОСКВА_ЛУЧИ", "Мавис": "МАВИС",
    "Мастер Недвижимость Елизаровский": "МАСТЕР_НЕДВИЖИМОСТЬ_ЕЛИЗАРОВСКИЙ",
    "Мастер Недвижимость Черная Речка": "МАСТЕР_НЕДВИЖИМОСТЬ_ЧЕРНАЯ", "МИЦ Бунино": "МИЦ_ЮЖНОЕ_БУНИНО",
    "Монолит Вингс": "МОНОЛИТ", "Ойкумена": "ОЙКУМЕНА",
    "Полис Групп Комендантский": "ПОЛИС_ГРУПП_КОМЕНДАНТСКИЙ", "Самолет": "САМОЛЕТ", "Севен Санс": "SEVEN_SUNS",
    "Технополис Премьера 2": "ТЕХНОПОЛИС_ПРЕМЬЕРА2", "ФСК Ай Проспект": "ФСК_АЙПРОСПЕКТ", "ЦДС": "ЦДС",
    "ЮИТ": "ЮИТ", "Технополис Рыбацкая Гавань": "ТЕХНОПОЛИС_РЫБАЦКАЯ",
    "Полис Групп Московский": "ПОЛИС_ГРУПП_МОСКОВСКИЙ", "ЛСР Зиларт": "ЛСР_МОСКВА_ЗИЛАРТ",
    "ЛСР Ленинградка": "ЛСР_МОСКВА_ЛЕНИНГРАДКА", "МИЦ ЭмДжиКом": "МИЦ_ЭМДЖИКОМ"
}


def importValues():
    wb = openpyxl.load_workbook('report.xlsx')
    ws = wb["Данные"]
    ws.insert_cols(1)
    ws["A1"] = "Клиент"
    # clientsCampaigns = {a: [] for a in clientList.keys()}
    for n in range(2, ws.max_row):
        print(ws["C"][n].value)
        for name in clientList.keys():
            if clientList.get(name) in str(ws["C"][n].value):
                _ = ws.cell(row=n + 1, column=1, value=name)
                print("Найден клиент %s для кампании %s" % (name, ws["C"][n].value))

    wb.save(filename="report.xlsx")


def sumClientPerformance():
    wb = openpyxl.load_workbook('report.xlsx')
    ws = wb["Данные"]

    # Сортируем лист по клиентам
    values = []
    for n, spreadsheetRow in enumerate(ws):
        if n > 0:
            values.append([])
            for spreadsheetCol in spreadsheetRow:
                if spreadsheetCol.value is not None:
                    if str(spreadsheetCol.value).isdigit():
                        values[n - 1].append(int(spreadsheetCol.value))
                    else:
                        values[n - 1].append(str(spreadsheetCol.value))
                else:
                    values[n - 1].append('')

    # Удаляем пустые строки (хз откуда они берутся)
    for x in values:
        c = 0
        for n in x:
            if n == '':
                c += 1
        if c >= 9:
            values.remove(x)
    tableValues = sorted(values, key=itemgetter(0))
    for n, row in enumerate(tableValues):
        for m, col in enumerate(row):
            _ = ws.cell(row=n + 2, column=m + 1, value=col)

    # Суммируем метрики клиентов
    clientMetrics = {}
    for client in clientList.keys():
        clientMetrics[client] = [0, 0, 0, 0]
        for row in range(2, ws.max_row):
            if clientList.get(client) in str(ws["C"][row].value):
                for col in range(5, 8):
                    clientMetrics[client][col - 5] += round(float(ws.cell(row=row, column=col).value), 0)
                if clientMetrics[client][3] != 0:
                    clientMetrics[client][4:5] += [clientMetrics[client][3] / clientMetrics[client][1],
                                                   clientMetrics[client][2] / clientMetrics[client][3]]
                print(clientMetrics)

    print(clientMetrics)
    # Вставляем строки с суммой

    wb.save(filename="report.xlsx")


# importValues()
sumClientPerformance()

print("\n--- %s seconds ---" % (time.time() - start_time))
