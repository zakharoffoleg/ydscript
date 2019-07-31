# -*- coding: utf-8 -*-
import io

from googleads import adwords
from openpyxl import Workbook


# billboadbot tokens
# access_token: ya29.GltVB1CZMI5EeKtudN4U1Q7UkyLkO0sowci47y3BsGXhPRVB7mvX3dN4QhMNdTjaEktaeP0VuSD91O3Q2wZYhnZ6egTr1uDFQ3OyoRVxy9tXBM6RbfOFedhi73OP
# refresh_token: 1/Jp1a51XGYcGLy3L_klNsDUU_iZhK6xd8-r3SARfIGEQ

# billboadbottest tokens
# Access token: ya29.GltVBwf6YydBqXGapPlRWzK9Ohw0l1pxEgGtWCuV2X88e6pcqE3PPrCzqJtNfqPcnx5RW3rBMO74-dS6DwdAIAVvDD2ZcwYwaJT-GEAdYTtw_loepaIV8xA1l_u_


class Account(object):
    def __init__(self, client):
        self.client = client

    def getCampaignCosts(self):
        # Initialize appropriate service.
        report_downloader = self.client.GetReportDownloader(version='v201809')

        # Create report query.
        report_query = (adwords.ReportQueryBuilder()
                        .Select('CampaignId', 'CampaignName', 'Cost')
                        .From(
            'CAMPAIGN_PERFORMANCE_REPORT')  # https://developers.google.com/adwords/api/docs/appendix/reports
                        # .Where('Status').In('ENABLED', 'PAUSED')
                        .During('LAST_7_DAYS')
                        .Build())

        # You can provide a file object to write the output to. For this
        # demonstration we use sys.stdout to write the report to the screen.

        with io.StringIO() as f:
            report_downloader.DownloadReportWithAwql(
                report_query, 'CSV', f, skip_report_header=False,
                skip_column_header=False, skip_report_summary=False,
                include_zero_impressions=True)
            # f.close()
            return f.getvalue()

    def exportToExcel(self):
        report = self.getCampaignCosts()
        wb = Workbook()
        ws = wb.active
        ws.title = "Отчёт GA"
        listReport = report.split(sep='\n')

        for num, col in enumerate(listReport[1].split(sep=',')):  # Вписываем заголовки таблицы
            _ = ws.cell(column=num + 1, row=1, value=col)

        for m, camp in enumerate(listReport[2:len(listReport) - 2]):
            campaign = camp.split(sep=',')
            for n, field in enumerate(campaign):
                if field.isdigit():
                    _ = ws.cell(column=n + 1, row=m + 2, value=int(field))
                else:
                    _ = ws.cell(column=n + 1, row=m + 2, value=field)
        wb.save(filename="report.xls")


if __name__ == '__main__':
    # Initialize client object.
    adwords_client = adwords.AdWordsClient.LoadFromStorage(path='googleads.yaml')
    User = Account(adwords_client)
    User.exportToExcel()
