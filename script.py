# -*- coding: utf-8 -*-
import collections
import io
import json
import re
import sys
from time import sleep

import requests
from googleads import adwords
from openpyxl import Workbook

from clients import callClients

if sys.version_info < (3,):
    def u(x):
        try:
            return x.encode("utf8")
        except UnicodeDecodeError:
            return x
else:
    def u(x):
        if type(x) == type(b''):
            return x.decode('utf8')
        else:
            return x

login = 'macsim.reshetar'
token = 'AgAAAAArGY4IAAXAIq_OsOIJXEgIvBjrf0dpP0s'  # Токен от macsim.reshetar
# token = 'AgAAAAAJq2baAAXKPEEN2051SEebhC-XYWWB8kQ'  # Токен от zakharoffffff
# token = 'AgAAAAAzN9ktAAXAInwzJfwlY0E4taHqCQUC33I' # Токен от za4aroff.ol


# billboadbot tokens
# access_token: ya29.GltVB1CZMI5EeKtudN4U1Q7UkyLkO0sowci47y3BsGXhPRVB7mvX3dN4QhMNdTjaEktaeP0VuSD91O3Q2wZYhnZ6egTr1uDFQ3OyoRVxy9tXBM6RbfOFedhi73OP
# refresh_token: 1/Jp1a51XGYcGLy3L_klNsDUU_iZhK6xd8-r3SARfIGEQ

# billboadbottest tokens
# Access token: ya29.GltVBwf6YydBqXGapPlRWzK9Ohw0l1pxEgGtWCuV2X88e6pcqE3PPrCzqJtNfqPcnx5RW3rBMO74-dS6DwdAIAVvDD2ZcwYwaJT-GEAdYTtw_loepaIV8xA1l_u_

wb = Workbook()
data = wb.active
data.title = 'Данные'
_ = data.cell(column=4, row=1, value='Рекламная система')


class YDAccount(object):

    def __init__(self, token, login):
        self.token = token
        self.login = login

    campaignList = []
    CampaignsURL = 'https://api.direct.yandex.com/json/v5/campaigns'
    ReportsURL = 'https://api.direct.yandex.com/json/v5/reports'

    def getCampaigns(self):
        headers = {"Authorization": "Bearer " + self.token,
                   "Client-Login": self.login,
                   "Accept-Language": "ru",
                   }

        body = {"method": "get",
                "params": {"SelectionCriteria": {},
                           "FieldNames": ["Id", "Name"]  # Отбирать только активные/остановленные кампании
                           }}

        jsonBody = json.dumps(body, ensure_ascii=False).encode('utf8')

        try:
            result = requests.post(self.CampaignsURL, jsonBody, headers=headers)

            # Отладочная информация
            # print("Заголовки запроса: {}".format(result.request.headers))
            # print("Запрос: {}".format(u(result.request.body)))
            # print("Заголовки ответа: {}".format(result.headers))
            # print("Ответ: {}".format(u(result.text)))

            # Обработка запроса
            if result.status_code != 200 or result.json().get("error", False):
                print("Произошла ошибка при обращении к серверу API Директа.")
                print("Код ошибки: {}".format(result.json()["error"]["error_code"]))
                print("Описание ошибки: {}".format(u(result.json()["error"]["error_detail"])))
                print("RequestId: {}".format(result.headers.get("RequestId", False)))
            else:
                print("RequestId: {}".format(result.headers.get("RequestId", False)))
                print("Информация о баллах: {}".format(result.headers.get("Units", False)))

                for campaign in result.json()["result"]["Campaigns"]:
                    self.campaignList.append(str(campaign['Id']))
                    # print("Рекламная кампания: {} №{}".format(u(campaign['Name']), campaign['Id']))
                    # print("%s, %s" % (campaign['Name'], campaign['Id']))

                if result.json()['result'].get('LimitedBy', False):
                    print("Получены не все доступные объекты.")

                return self.campaignList

        except ConnectionError:
            print("Произошла ошибка соединения с сервером API.")

        except:
            print("Произошла непредвиденная ошибка.")

    def getCosts(self):
        self.getCampaigns()
        if self.campaignList:
            self.campaignCosts = {}
            for i in self.campaignList:
                headers = {
                    "Authorization": "Bearer " + self.token,
                    "Client-Login": self.login,
                    "Accept-Language": "ru",
                    "processingMode": "auto"
                    # Формат денежных значений в отчете
                    # "returnMoneyInMicros": "false",
                    # Не выводить в отчете строку с названием отчета и диапазоном дат
                    # "skipReportHeader": "true",
                    # Не выводить в отчете строку с названиями полей
                    # "skipColumnHeader": "true",
                    # Не выводить в отчете строку с количеством строк статистики
                    # "skipReportSummary": "true"
                }

                self.costsBody = {
                    "params": {
                        "SelectionCriteria": {
                            "DateFrom": "2019-07-22",
                            "DateTo": "2019-07-23",
                            "Filter": [{
                                "Field": "CampaignId",
                                "Operator": "EQUALS",
                                "Values": [i],
                            }],
                        },
                        "FieldNames": [
                            # "Date",
                            "CampaignId",
                            "CampaignName",
                            # "LocationOfPresenceName",
                            # "Impressions",
                            # "Clicks",
                            "Cost"
                        ],
                        "ReportName": u("НАЗВАНИЕ_ОТЧЕТА"),
                        "ReportType": "CUSTOM_REPORT",
                        "DateRangeType": "CUSTOM_DATE",
                        "Format": "TSV",
                        "IncludeVAT": "YES",
                        "IncludeDiscount": "NO"
                    }
                }

                # Кодирование тела запроса в JSON
                body = json.dumps(self.costsBody, indent=4)

                # --- Запуск цикла для выполнения запросов ---
                # Если получен HTTP-код 200, то выводится содержание отчета
                # Если получен HTTP-код 201 или 202, выполняются повторные запросы
                while True:
                    try:
                        req = requests.post(self.ReportsURL, body, headers=headers)
                        req.encoding = 'utf-8'  # Принудительная обработка ответа в кодировке UTF-8
                        if req.status_code == 400:
                            print("Параметры запроса указаны неверно или достигнут лимит отчетов в очереди")
                            print("RequestId: {}".format(req.headers.get("RequestId", False)))
                            print("JSON-код запроса: {}".format(u(body)))
                            print("JSON-код ответа сервера: \n{}".format(u(req.json())))
                            break
                        elif req.status_code == 200:
                            # print("RequestId: {}".format(req.headers.get("RequestId", False)))
                            # print("Содержание отчета: \n{}".format(u(req.text)))
                            self.campaignCosts.update({i: req.text})
                            print("Данные о расходах кампании %s получены" % i)
                            break
                        elif req.status_code == 201:
                            print("Отчет успешно поставлен в очередь в режиме офлайн")
                            retryIn = int(req.headers.get("retryIn", 60))
                            print("Повторная отправка запроса через {} секунд".format(retryIn))
                            print("RequestId: {}".format(req.headers.get("RequestId", False)))
                            sleep(retryIn)
                        elif req.status_code == 202:
                            print("Отчет формируется в режиме офлайн")
                            retryIn = int(req.headers.get("retryIn", 60))
                            print("Повторная отправка запроса через {} секунд".format(retryIn))
                            print("RequestId:  {}".format(req.headers.get("RequestId", False)))
                            sleep(retryIn)
                        elif req.status_code == 500:
                            print("При формировании отчета произошла ошибка. Пожалуйста, попробуйте повторить запрос "
                                  "позднее")
                            print("RequestId: {}".format(req.headers.get("RequestId", False)))
                            print("JSON-код ответа сервера: \n{}".format(u(req.json())))
                            break
                        elif req.status_code == 502:
                            print("Время формирования отчета превысило серверное ограничение.")
                            print("Пожалуйста, попробуйте изменить параметры запроса - уменьшить период и количество "
                                  "запрашиваемых данных.")
                            print("JSON-код запроса: {}".format(body))
                            print("RequestId: {}".format(req.headers.get("RequestId", False)))
                            print("JSON-код ответа сервера: \n{}".format(u(req.json())))
                            break
                        else:
                            print("Произошла непредвиденная ошибка")
                            print("RequestId:  {}".format(req.headers.get("RequestId", False)))
                            print("JSON-код запроса: {}".format(body))
                            print("JSON-код ответа сервера: \n{}".format(u(req.json())))
                            break

                    # Обработка ошибки, если не удалось соединиться с сервером API Директа
                    except requests.exceptions.ConnectionError:
                        # В данном случае мы рекомендуем повторить запрос позднее
                        print("Произошла ошибка соединения с сервером API")
                        # Принудительный выход из цикла
                        break

                    # Если возникла какая-либо другая ошибка
                    except requests.exceptions.RequestException as e:
                        # В данном случае мы рекомендуем проанализировать действия приложения
                        print(e)
                        print("Произошла непредвиденная ошибка")
                        # Принудительный выход из цикла
                        break
            self.campaignCosts = collections.OrderedDict(sorted(self.campaignCosts.items()))
            return self.campaignCosts
        else:
            print("Список кампаний пуст!")

    def exportToExcel(self):
        self.getCosts()
        # wb = Workbook()
        # ws = wb.active
        # ws.title = "Отчёт YD"
        for col in range(len(self.costsBody.get("params").get("FieldNames"))):  # Проставляем графы таблицы
            _ = data.cell(column=col + 1, row=1, value=self.costsBody.get("params").get("FieldNames")[col])

        rowNum, colNum = 2, 1  # Содержание отчёта начинается с 73-го символа
        for i in self.campaignCosts.items():
            rows = [a.start() for a in list(re.finditer('\n', i[1]))]
            cols = [a.start() for a in list(re.finditer('\t', i[1]))]
            if rows and cols:
                index = 73
                x = 0
                while x < len(rows) and x < len(cols):
                    if rows[x] < index:
                        del rows[x]
                    elif cols[x] < index:
                        del cols[x]
                    else:
                        x += 1

                for n in range(len(rows) + len(cols)):
                    if not rows:
                        for col in cols:
                            if i[1][index:cols[0]].strip():
                                if i[1][index:cols[0]].strip().isdigit():
                                    _ = data.cell(column=colNum, row=rowNum, value=int(i[1][index:cols[0]].strip())
                                                                                   / 10 ** 6)
                                else:
                                    _ = data.cell(column=colNum, row=rowNum, value=i[1][index:cols[0]].strip())
                                index = cols[0]
                                cols.pop(0)
                                colNum += 1
                    elif not cols:
                        if len(rows) >= 2:
                            for row in rows:
                                if i[1][index:rows[0]].strip():
                                    if i[1][index:rows[0]].strip().isdigit():
                                        _ = data.cell(column=colNum, row=rowNum, value=int(i[1][index:rows[0]].strip())
                                                                                       / 10 ** 6)
                                    else:
                                        _ = data.cell(column=colNum, row=rowNum, value=i[1][index:rows[0]].strip())
                                    index = rows[0]
                                    rows.pop(0)
                                    rowNum += 1
                                    colNum = 1
                        else:
                            print("Отчёт по кампании %s сформирован!" % (i[0]))
                    else:
                        if rows[0] < cols[0]:
                            if i[1][index:rows[0]].strip():
                                if i[1][index:rows[0]].strip().isdigit():
                                    _ = data.cell(column=colNum, row=rowNum, value=int(i[1][index:rows[0]].strip())
                                                                                   / 10 ** 6)
                                else:
                                    _ = data.cell(column=colNum, row=rowNum, value=i[1][index:rows[0]].strip())
                                index = rows[0]
                                rows.pop(0)
                                rowNum += 1
                                colNum = 1
                        else:
                            if i[1][index:cols[0]].strip():
                                if i[1][index:cols[0]].strip().isdigit():
                                    _ = data.cell(column=colNum, row=rowNum, value=int(i[1][index:cols[0]].strip())
                                                                                   / 10 ** 6)
                                else:
                                    _ = data.cell(column=colNum, row=rowNum, value=i[1][index:cols[0]].strip())
                                index = cols[0]
                                cols.pop(0)
                                colNum += 1

            else:
                print("Нет списка расходов по кампаниям!")
            colNum = 1
            _ = data.cell(column=4, row=rowNum, value='YD')
        # wb.save(filename="report.xls")


class GAAccount(object):

    def __init__(self, client):
        self.client = client

    def getCampaignCosts(self):
        # Initialize appropriate service.
        report_downloader = self.client.GetReportDownloader(version='v201809')

        # Create report query.
        report_query = (adwords.ReportQueryBuilder()
                        .Select('CampaignId', 'CampaignName', 'Cost')
                        .From(
            'CAMPAIGN_PERFORMANCE_REPORT')  # https://developers.google.com/adwords/api/docs/appendix/reports
                        # .Where('Status').In('ENABLED', 'PAUSED')
                        .During('LAST_7_DAYS')
                        .Build())

        # You can provide a file object to write the output to. For this
        # demonstration we use sys.stdout to write the report to the screen.

        with io.StringIO() as f:
            report_downloader.DownloadReportWithAwql(
                report_query, 'CSV', f, skip_report_header=False,
                skip_column_header=False, skip_report_summary=False,
                include_zero_impressions=True)
            # f.close()
            return f.getvalue()

    def exportToExcel(self):
        report = self.getCampaignCosts()
        # gasheet = wb.create_sheet(title='Отчёт GA')
        # wb = Workbook()
        # ws = wb.active
        # ws.title = "Отчёт GA"
        listReport = report.split(sep='\n')

        # for num, col in enumerate(listReport[1].split(sep=',')):  # Вписываем заголовки таблицы
        #    _ = gasheet.cell(column=num + 1, row=1, value=col)

        for m, camp in enumerate(listReport[2:len(listReport) - 2]):
            campaign = camp.split(sep=',')
            for n, field in enumerate(campaign):
                if field.isdigit():
                    _ = data.cell(column=n + 1, row=data.max_row + m,
                                  value=int(field))  # ws.max_row - потенциально косяк, если кампаний больше 1
                else:
                    _ = data.cell(column=n + 1, row=data.max_row + m, value=field)
            _ = data.cell(column=4, row=data.max_row + m, value='GA')
        # wb.save(filename="report.xls")


YDUser = YDAccount(token, login)
# YDUser.exportToExcel()

adwords_client = adwords.AdWordsClient.LoadFromStorage(path='googleads.yaml')
GAUser = GAAccount(adwords_client)


# GAUser.exportToExcel()


def exportCosts():
    YDUser.exportToExcel()
    GAUser.exportToExcel()


def sumCosts():
    report = wb.create_sheet('Отчёт', 0)
    # report = wb.active
    for n, client in enumerate(callClients):
        _ = report.cell(column=1, row=n + 1, value=client)
    wb.save(filename="report.xls")


exportCosts()
sumCosts()
