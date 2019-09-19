# -*- coding: utf-8 -*-
import io
import json
import sys
import time
from time import sleep

import requests
from googleads import adwords
from openpyxl import Workbook

start_time = time.time()

if sys.version_info < (3,):
    def u(x):
        try:
            return x.encode("utf8")
        except UnicodeDecodeError:
            return x
else:
    def u(x):
        if type(x) == type(b''):
            return x.decode('utf8')
        else:
            return x

# Auth info
YDAuth = {'za4aroff.ol': 'AgAAAAAzN9ktAAXAInwzJfwlY0E4taHqCQUC33I',
          'macsim.reshetar': 'AgAAAAArGY4IAAXAIq_OsOIJXEgIvBjrf0dpP0s'}


# token = 'AgAAAAAzN9ktAAXAInwzJfwlY0E4taHqCQUC33I' # Токен от za4aroff.ol


# billboadbot tokens
# access_token: ya29.GltVB1CZMI5EeKtudN4U1Q7UkyLkO0sowci47y3BsGXhPRVB7mvX3dN4QhMNdTjaEktaeP0VuSD91O3Q2wZYhnZ6egTr1uDFQ3OyoRVxy9tXBM6RbfOFedhi73OP
# refresh_token: 1/Jp1a51XGYcGLy3L_klNsDUU_iZhK6xd8-r3SARfIGEQ

# billboadbottest tokens
# Access token: ya29.GltVBwf6YydBqXGapPlRWzK9Ohw0l1pxEgGtWCuV2X88e6pcqE3PPrCzqJtNfqPcnx5RW3rBMO74-dS6DwdAIAVvDD2ZcwYwaJT-GEAdYTtw_loepaIV8xA1l_u_

wb = Workbook()
data = wb.active
data.title = 'Данные'
_ = data.cell(column=4, row=1, value='Рекламная система')


class YDAccount(object):

    def __init__(self, token, login):
        self.token = token
        self.login = login

    campaignList = []
    CampaignsURL = 'https://api.direct.yandex.com/json/v5/campaigns'
    ReportsURL = 'https://api.direct.yandex.com/json/v5/reports'

    def getCosts(self):
        self.responseReport = ''
        headers = {
            "Authorization": "Bearer " + self.token,
            "Client-Login": self.login,
            "Accept-Language": "ru",
            "processingMode": "auto",
            # Формат денежных значений в отчете
            "returnMoneyInMicros": "false",
            # Не выводить в отчете строку с названием отчета и диапазоном дат
            # "skipReportHeader": "true",
            # Не выводить в отчете строку с названиями полей
            # "skipColumnHeader": "true",
            # Не выводить в отчете строку с количеством строк статистики
            # "skipReportSummary": "true"
        }
        self.costsBody = {
            "params": {
                "SelectionCriteria": {
                    "DateFrom": "2019-09-01",  # Первый день месяца
                    "DateTo": "2019-09-18",  # Вчера
                    "Filter": [{
                        "Field": "Impressions",
                        "Operator": "GREATER_THAN",
                        "Values": [0],
                    }],
                },
                "FieldNames": [
                    # "Date",
                    "CampaignId",
                    "CampaignName",
                    "AdGroupName",
                    # "LocationOfPresenceName",
                    "Impressions",
                    "Clicks",
                    "Cost"
                ],
                "ReportName": u("НАЗВAНИЕ_OТЧЕТA"),
                "ReportType": "CUSTOM_REPORT",
                "DateRangeType": "CUSTOM_DATE",
                "Format": "TSV",
                "IncludeVAT": "YES",
                "IncludeDiscount": "NO"
            }
        }

        # Кодирование тела запроса в JSON
        body = json.dumps(self.costsBody, indent=4)

        # --- Запуск цикла для выполнения запросов ---
        # Если получен HTTP-код 200, то выводится содержание отчета
        # Если получен HTTP-код 201 или 202, выполняются повторные запросы
        while True:
            try:
                req = requests.post(self.ReportsURL, body, headers=headers)
                req.encoding = 'utf-8'  # Принудительная обработка ответа в кодировке UTF-8
                # print(req.text)
                if req.status_code == 400:
                    print("Параметры запроса указаны неверно или достигнут лимит отчетов в очереди")
                    print("RequestId: {}".format(req.headers.get("RequestId", False)))
                    print("JSON-код запроса: {}".format(u(body)))
                    print("JSON-код ответа сервера: \n{}".format(u(req.json())))
                    break
                elif req.status_code == 200:
                    # print("RequestId: {}".format(req.headers.get("RequestId", False)))
                    # print("Содержание отчета: \n{}".format(u(req.text)))
                    self.responseReport += req.text
                    print("Данные о расходах кампаний получены")
                    break
                elif req.status_code == 201:
                    print("Отчет успешно поставлен в очередь в режиме офлайн")
                    retryIn = int(req.headers.get("retryIn", 60))
                    print("Повторная отправка запроса через {} секунд".format(retryIn))
                    print("RequestId: {}".format(req.headers.get("RequestId", False)))
                    sleep(retryIn)
                elif req.status_code == 202:
                    print("Отчет формируется в режиме офлайн")
                    retryIn = int(req.headers.get("retryIn", 60))
                    print("Повторная отправка запроса через {} секунд".format(retryIn))
                    print("RequestId:  {}".format(req.headers.get("RequestId", False)))
                    sleep(retryIn)
                elif req.status_code == 500:
                    print("При формировании отчета произошла ошибка. Пожалуйста, попробуйте повторить запрос "
                          "позднее")
                    print("RequestId: {}".format(req.headers.get("RequestId", False)))
                    print("JSON-код ответа сервера: \n{}".format(u(req.json())))
                    break
                elif req.status_code == 502:
                    print("Время формирования отчета превысило серверное ограничение.")
                    print("Пожалуйста, попробуйте изменить параметры запроса - уменьшить период и количество "
                          "запрашиваемых данных.")
                    print("JSON-код запроса: {}".format(body))
                    print("RequestId: {}".format(req.headers.get("RequestId", False)))
                    print("JSON-код ответа сервера: \n{}".format(u(req.json())))
                    break
                else:
                    print("Произошла непредвиденная ошибка")
                    print("RequestId:  {}".format(req.headers.get("RequestId", False)))
                    print("JSON-код запроса: {}".format(body))
                    print("JSON-код ответа сервера: \n{}".format(u(req.json())))
                    break

                    # Обработка ошибки, если не удалось соединиться с сервером API Директа
            except requests.exceptions.ConnectionError:
                # В данном случае мы рекомендуем повторить запрос позднее
                print("Произошла ошибка соединения с сервером API")
                # Принудительный выход из цикла
                break

                # Если возникла какая-либо другая ошибка
            except requests.exceptions.RequestException as e:
                # В данном случае мы рекомендуем проанализировать действия приложения
                print(e)
                print("Произошла непредвиденная ошибка")
                # Принудительный выход из цикла
                break
        return self.responseReport

    def exportToExcel(self):
        self.getCosts()
        for col in range(len(self.costsBody.get("params").get("FieldNames"))):  # Проставляем графы таблицы
            _ = data.cell(column=col + 1, row=1, value=self.costsBody.get("params").get("FieldNames")[col])

        rows = self.responseReport.split('\n')
        maxRow = data.max_row + 1
        for m, camp in enumerate(rows[2:len(rows) - 2]):
            campaign = camp.split(sep='\t')
            print('Данные кампании %s получены.' % campaign[0])
            for n, field in enumerate(campaign):
                if field.isdigit():
                    _ = data.cell(column=n + 1, row=maxRow, value=float(field.replace('.', ',')))
                else:
                    _ = data.cell(column=n + 1, row=maxRow, value=field)
            maxRow += 1
        wb.save(filename="report.xlsx")
        return


class GAAccount(object):

    def __init__(self, client):
        self.client = client

    def getCampaignCosts(self):
        # Initialize appropriate service.
        report_downloader = self.client.GetReportDownloader(version='v201809')

        # Create report query.
        report_query = (adwords.ReportQueryBuilder()
                        .Select('CampaignId', 'CampaignName', 'AdGroupId', 'AdGroupName' 'Cost')
                        .From(
            'ADGROUP_PERFORMANCE_REPORT')  # https://developers.google.com/adwords/api/docs/appendix/reports
                        # .Where('Status').In('ENABLED', 'PAUSED')
                        .During('LAST_7_DAYS')
                        .Build())

        # You can provide a file object to write the output to. For this
        # demonstration we use sys.stdout to write the report to the screen.

        with io.StringIO() as f:
            report_downloader.DownloadReportWithAwql(
                report_query, 'CSV', f, skip_report_header=False,
                skip_column_header=False, skip_report_summary=False,
                include_zero_impressions=True)
            # f.close()
            return f.getvalue()

    def exportToExcel(self):
        report = self.getCampaignCosts()
        listedReport = report.split(sep='\n')

        maxRow = data.max_row + 1
        for m, camp in enumerate(listedReport[2:len(listedReport) - 2]):
            campaign = camp.split(sep=',')
            for n, field in enumerate(campaign):
                if field.isdigit():
                    _ = data.cell(column=n + 1, row=maxRow, value=float(field))
                else:
                    _ = data.cell(column=n + 1, row=maxRow, value=field)
            maxRow += 1
        wb.save(filename="report.xlsx")


adwords_client = adwords.AdWordsClient.LoadFromStorage(path='googleads.yaml')
GAUser = GAAccount(adwords_client)


def getCosts():
    for i in YDAuth.items():
        YDUser = YDAccount(i[1], i[0])
        YDUser.exportToExcel()
    # GAUser.exportToExcel()


def sumCosts():  # Может сломаться из-за новых полей показов и кликов
    # wb = openpyxl.load_workbook('C:/Users/KK/PycharmProjects/ydscript/report.xlsx')
    # report = wb.active
    report = wb.create_sheet('Отчёт', 0)
    for n, field in enumerate(['Кампания', 'Группа', 'Расход', 'Конверсии', '% конверсии', 'CPA']):  # Вводим заголовки
        _ = report.cell(column=1 + n, row=1, value=field)
    # for n, client in enumerate(callClients):  # Вводим список клиентов
    #    _ = report.cell(column=1, row=n + 2, value=client)

    for clientRow, i in enumerate(report['A'][1:]):
        print(i.value)

        autosum = 0
        for row, x in enumerate(wb['Данные']['B'][1:]):
            if i.value.upper() in x.value:
                print(x.value)
                autosum += float(wb['Данные']['C'][row + 1].value)
        # print(autosum)
        _ = report.cell(column=2, row=clientRow + 2, value=autosum)

    wb.save(filename="report.xlsx")


getCosts()
#sumCosts()

print("\n--- %s seconds ---" % (time.time() - start_time))
